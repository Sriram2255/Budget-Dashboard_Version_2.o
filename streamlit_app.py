import os
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, date
import sqlite3
import uuid
import random
import string
from pathlib import Path

# Configure page
st.set_page_config(
    page_title='Comprehensive Project Management Dashboard', 
    layout='wide',
    initial_sidebar_state='expanded',
    page_icon='🚀'
)

# No auto-open to prevent multiple browser windows

# Initialize Streamlit session state defaults to avoid KeyError on first load
if "project_info_submitted" not in st.session_state:
    st.session_state["project_info_submitted"] = False
if "project_basic" not in st.session_state:
    st.session_state["project_basic"] = {}
if "project_materials_data" not in st.session_state:
    st.session_state["project_materials_data"] = []
if "materials_data" not in st.session_state:
    st.session_state["materials_data"] = []
if "selected_category_filter" not in st.session_state:
    st.session_state["selected_category_filter"] = "All Categories"

# Initialize database
def init_database():
    conn = sqlite3.connect('project_management.db')
    cursor = conn.cursor()
    
    # Create projects table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tracking_id TEXT UNIQUE,
            project_name TEXT NOT NULL,
            project_description TEXT,
            domain TEXT,
            priority TEXT,
            estimated_budget REAL,
            manpower_count INTEGER,
            manpower_cost REAL,
            material_cost REAL,
            equipment_cost REAL,
            other_costs REAL,
            total_cost REAL,
            start_date DATE,
            end_date DATE,
            department TEXT,
            contact_email TEXT,
            contact_phone TEXT,
            justification TEXT,
            risk_assessment TEXT,
            expected_outcome TEXT,
            status TEXT DEFAULT 'pending',
            submitted_by TEXT,
            submitted_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            reviewed_by TEXT,
            review_date TIMESTAMP,
            review_comments TEXT
        )
    ''')
    
    # Create notifications table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS notifications (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            notification_type TEXT,
            title TEXT,
            message TEXT,
            is_read BOOLEAN DEFAULT FALSE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY (project_id) REFERENCES projects (id)
        )
    ''')
    # Create project_materials table if not exists (preserve existing data)
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS project_materials (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            project_id INTEGER,
            category TEXT,
            subtopic TEXT,
            description TEXT,
            units_qty TEXT,
            nos INTEGER,
            source_type TEXT,
            payment_schedule TEXT,
            unit_price REAL,
            amount_inr REAL,
            justification TEXT,
            justification_type TEXT,
            justification_file_path TEXT,
            status TEXT DEFAULT 'pending',
            review_comments TEXT,
            reviewed_by TEXT,
            reviewed_at TIMESTAMP,
            finalized BOOLEAN DEFAULT 0,
            FOREIGN KEY (project_id) REFERENCES projects (id)
        )
    ''')

    # Lightweight migration: add missing columns if the table already exists
    cursor.execute("PRAGMA table_info(project_materials)")
    existing_cols = {row[1] for row in cursor.fetchall()}
    columns_to_add = [
        ("status", "TEXT", "'pending'"),
        ("review_comments", "TEXT", None),
        ("reviewed_by", "TEXT", None),
        ("reviewed_at", "TIMESTAMP", None),
        ("finalized", "BOOLEAN", "0"),
    ]
    for col_name, col_type, default_val in columns_to_add:
        if col_name not in existing_cols:
            if default_val is not None:
                cursor.execute(f"ALTER TABLE project_materials ADD COLUMN {col_name} {col_type} DEFAULT {default_val}")
            else:
                cursor.execute(f"ALTER TABLE project_materials ADD COLUMN {col_name} {col_type}")
    
    conn.commit()
    conn.close()

# Generate tracking ID
def generate_tracking_id():
    prefix = "PRJ"
    timestamp = datetime.now().strftime("%y%m%d")
    random_suffix = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
    return f"{prefix}{timestamp}{random_suffix}"

# Database operations
def get_projects():
    conn = sqlite3.connect('project_management.db')
    df = pd.read_sql_query("SELECT * FROM projects ORDER BY submitted_at DESC", conn)
    conn.close()
    return df

def get_project_by_tracking_id(tracking_id):
    conn = sqlite3.connect('project_management.db')
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM projects WHERE tracking_id = ?", (tracking_id,))
    result = cursor.fetchone()
    conn.close()
    return result

def save_project(project_data):
    conn = sqlite3.connect('project_management.db')
    cursor = conn.cursor()
    
    tracking_id = generate_tracking_id()
    
    # Calculate cumulative costs from materials
    materials_data = st.session_state.get("project_materials_data", [])
    material_cost = sum(item.get("Amount INR", 0) for item in materials_data)
    
    # Set default values for simplified form
    estimated_budget = material_cost  # Auto-calculated from materials
    manpower_count = 0
    manpower_cost = 0
    equipment_cost = 0
    other_costs = 0
    total_cost = material_cost
    
    # Safely get optional fields
    risk_assessment = project_data.get('risk_assessment', '')
    expected_outcome = project_data.get('expected_outcome', '')
    
    cursor.execute('''
        INSERT INTO projects (
            tracking_id, project_name, project_description, domain, priority,
            estimated_budget, manpower_count, manpower_cost, material_cost,
            equipment_cost, other_costs, total_cost, start_date, end_date,
            department, contact_email, contact_phone, justification,
            risk_assessment, expected_outcome, submitted_by
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        tracking_id, project_data['project_name'], project_data['project_description'],
        project_data['domain'], project_data['priority'], estimated_budget,
        manpower_count, manpower_cost, material_cost,
        equipment_cost, other_costs, total_cost,
        project_data['start_date'], project_data['end_date'], project_data['department'],
        project_data['contact_email'], project_data['contact_phone'], project_data['justification'],
        risk_assessment, expected_outcome, project_data['submitted_by']
    ))
    
    conn.commit()
    conn.close()
    return tracking_id

def update_project_status(tracking_id, status, review_comments, reviewed_by):
    conn = sqlite3.connect('project_management.db')
    cursor = conn.cursor()
    
    cursor.execute('''
        UPDATE projects 
        SET status = ?, review_comments = ?, reviewed_by = ?, review_date = CURRENT_TIMESTAMP
        WHERE tracking_id = ?
    ''', (status, review_comments, reviewed_by, tracking_id))
    
    conn.commit()
    conn.close()

def get_notifications():
    conn = sqlite3.connect('project_management.db')
    df = pd.read_sql_query("SELECT * FROM notifications ORDER BY created_at DESC", conn)
    conn.close()
    return df

def add_notification(project_id, notification_type, title, message):
    conn = sqlite3.connect('project_management.db')
    cursor = conn.cursor()
    
    cursor.execute('''
        INSERT INTO notifications (project_id, notification_type, title, message)
        VALUES (?, ?, ?, ?)
    ''', (project_id, notification_type, title, message))
    
    conn.commit()
    conn.close()

# Update the save_materials function to include category, subtopic, and justification
def save_materials(project_id, materials):
    conn = sqlite3.connect('project_management.db')
    cursor = conn.cursor()
    for entry in materials:
        cursor.execute('''
            INSERT INTO project_materials (
                project_id, category, subtopic, description, units_qty, nos, source_type, payment_schedule, unit_price, amount_inr,
                justification, justification_type, justification_file_path, status, finalized
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            project_id,
            entry['Category'],
            entry['Sub-topic'],
            entry['Description'],
            entry['Units/Qty'],
            entry['Nos'],
            entry['Source/Type'],
            entry['Payment Schedule'],
            entry['Unit Price Total Amount'],
            entry['Amount INR'],
            entry.get('Justification', ''),
            entry.get('Justification Type', ''),
            entry.get('Justification File Path', ''),
            'pending',
            0
        ))
    conn.commit()
    conn.close()

def get_materials_by_project(project_id):
    conn = sqlite3.connect('project_management.db')
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM project_materials WHERE project_id = ?', (project_id,))
    data = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]
    df = pd.DataFrame(data, columns=columns)
    conn.close()
    return df

def update_material_status(material_id, status, review_comments=None, reviewed_by='Admin', finalize=False):
    conn = sqlite3.connect('project_management.db')
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE project_materials
        SET status = ?, review_comments = ?, reviewed_by = ?, reviewed_at = CURRENT_TIMESTAMP, finalized = ?
        WHERE id = ?
    ''', (status, review_comments, reviewed_by, 1 if finalize else 0, material_id))
    conn.commit()
    conn.close()

# Initialize database
init_database()

# Professional CSS Styling
st.markdown("""
    <style>
    /* Main Theme Colors */
    :root {
        --primary-color: #1f77b4;
        --secondary-color: #ff7f0e;
        --success-color: #2ca02c;
        --warning-color: #d62728;
        --info-color: #17a2b8;
        --light-bg: #f8f9fa;
        --dark-text: #2c3e50;
        --border-color: #dee2e6;
    }

    /* Main Container */
    .main-header {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 2rem 1rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        color: white;
        text-align: center;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
    }

    .main-header h1 {
        font-size: 2.5rem;
        font-weight: 700;
        margin: 0;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
    }

    .main-header p {
        font-size: 1.2rem;
        margin: 0.5rem 0 0 0;
        opacity: 0.9;
    }

    /* Cards */
    .metric-card {
        background: white;
        padding: 1.5rem;
        border-radius: 12px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.08);
        border-left: 4px solid var(--primary-color);
        transition: transform 0.2s ease;
    }

    .metric-card:hover {
        transform: translateY(-2px);
        box-shadow: 0 4px 20px rgba(0,0,0,0.12);
    }

    .metric-value {
        font-size: 2rem;
        font-weight: 700;
        color: var(--primary-color);
        margin: 0;
    }

    .metric-label {
        font-size: 0.9rem;
        color: #6c757d;
        margin: 0.5rem 0 0 0;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }

    /* Status Badges */
    .status-badge {
        padding: 0.5rem 1rem;
        border-radius: 20px;
        font-weight: 600;
        font-size: 0.8rem;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }

    .status-pending {
        background-color: #fff3cd;
        color: #856404;
        border: 1px solid #ffeaa7;
    }

    .status-approved {
        background-color: #d4edda;
        color: #155724;
        border: 1px solid #c3e6cb;
    }

    .status-rejected {
        background-color: #f8d7da;
        color: #721c24;
        border: 1px solid #f5c6cb;
    }

    .status-under-review {
        background-color: #d1ecf1;
        color: #0c5460;
        border: 1px solid #bee5eb;
    }

    /* Forms */
    .form-container {
        background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
        padding: 2rem;
        border-radius: 12px;
        box-shadow: 0 4px 20px rgba(0,0,0,0.1);
        margin-bottom: 2rem;
        border: 1px solid #dee2e6;
    }

    .form-section {
        background: rgba(255, 255, 255, 0.8);
        padding: 1.5rem;
        margin-bottom: 1.5rem;
        border-radius: 8px;
        border-bottom: 2px solid #e9ecef;
        border-left: 4px solid var(--primary-color);
    }

    .topic-section {
        background: rgba(255, 255, 255, 0.9);
        padding: 1rem;
        margin-bottom: 1rem;
        border-radius: 6px;
        border-bottom: 1px solid #dee2e6;
        border-left: 3px solid var(--info-color);
    }

    .topic-section h3 {
        margin-bottom: 0.5rem;
        padding-bottom: 0.5rem;
        border-bottom: 2px solid var(--primary-color);
        color: var(--primary-color);
        font-weight: 600;
    }

    /* Buttons */
    .stButton > button {
        background: linear-gradient(45deg, var(--primary-color), #4a90e2);
        color: white;
        border: none;
        border-radius: 8px;
        padding: 0.5rem 1.5rem;
        font-weight: 600;
        transition: all 0.3s ease;
        box-shadow: 0 2px 8px rgba(0,0,0,0.15);
    }

    .stButton > button:hover {
        transform: translateY(-1px);
        box-shadow: 0 4px 15px rgba(0,0,0,0.2);
    }

    /* Sidebar */
    .sidebar .sidebar-content {
        background: linear-gradient(180deg, #f8f9fa 0%, #e9ecef 100%);
    }

    /* Progress Bars */
    .progress-container {
        background: #e9ecef;
        border-radius: 10px;
        padding: 0.5rem;
        margin: 1rem 0;
    }

    .progress-bar {
        background: linear-gradient(90deg, var(--success-color), #4caf50);
        height: 8px;
        border-radius: 5px;
        transition: width 0.3s ease;
    }

    /* Tables */
    .dataframe {
        border-radius: 8px;
        overflow: hidden;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
        background: rgba(255, 255, 255, 0.95);
        border: 1px solid #dee2e6;
    }

    /* Data display improvements */
    .stDataFrame {
        background: rgba(255, 255, 255, 0.95);
        border-radius: 8px;
        border: 1px solid #dee2e6;
    }

    /* Enhanced table styling */
    .stDataFrame table {
        border-collapse: collapse;
        width: 100%;
        font-size: 14px;
    }

    .stDataFrame th {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        font-weight: 600;
        padding: 12px 8px;
        text-align: left;
        border-bottom: 2px solid #5a67d8;
    }

    .stDataFrame td {
        padding: 10px 8px;
        border-bottom: 1px solid #e2e8f0;
        background: rgba(255, 255, 255, 0.9);
    }

    .stDataFrame tr:nth-child(even) td {
        background: rgba(248, 250, 252, 0.9);
    }

    .stDataFrame tr:hover td {
        background: rgba(237, 242, 247, 0.9);
        transition: background-color 0.2s ease;
    }

    /* BOQ Review Table Styling */
    .boq-review-table {
        background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
        border-radius: 12px;
        padding: 1.5rem;
        margin: 1rem 0;
        border: 2px solid #dee2e6;
        box-shadow: 0 6px 25px rgba(0,0,0,0.1);
    }

    .boq-category-header {
        background: linear-gradient(135deg, #17a2b8 0%, #138496 100%);
        color: white;
        padding: 1rem;
        border-radius: 8px;
        margin-bottom: 1rem;
        font-weight: 600;
        font-size: 1.1rem;
        box-shadow: 0 2px 10px rgba(0,0,0,0.1);
    }

    /* Excel-like Table Styling */
    .excel-table-container {
        background: white;
        border: 2px solid #d1d5db;
        border-radius: 8px;
        overflow: hidden;
        box-shadow: 0 4px 20px rgba(0,0,0,0.1);
        margin: 1rem 0;
    }

    .excel-table-container .stDataFrame {
        border: none;
        background: white;
    }

    .excel-table-container .stDataFrame table {
        border-collapse: collapse;
        width: 100%;
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
        font-size: 13px;
    }

    .excel-table-container .stDataFrame th {
        background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
        color: #2c3e50;
        font-weight: 700;
        padding: 12px 8px;
        text-align: left;
        border: 1px solid #dee2e6;
        border-bottom: 2px solid #adb5bd;
        font-size: 12px;
        text-transform: uppercase;
        letter-spacing: 0.5px;
    }

    .excel-table-container .stDataFrame td {
        padding: 10px 8px;
        border: 1px solid #dee2e6;
        background: white;
        font-size: 13px;
        vertical-align: top;
    }

    .excel-table-container .stDataFrame tr:nth-child(even) td {
        background: #f8f9fa;
    }

    .excel-table-container .stDataFrame tr:hover td {
        background: #e3f2fd;
        transition: background-color 0.2s ease;
    }

    /* Status-specific row coloring */
    .excel-table-container .stDataFrame tr[data-status="approved"] td {
        background: #d4edda !important;
        border-left: 4px solid #28a745;
    }

    .excel-table-container .stDataFrame tr[data-status="pending"] td {
        background: #fff3cd !important;
        border-left: 4px solid #ffc107;
    }

    .excel-table-container .stDataFrame tr[data-status="under_review"] td {
        background: #d1ecf1 !important;
        border-left: 4px solid #17a2b8;
    }

    .excel-table-container .stDataFrame tr[data-status="rejected"] td {
        background: #f8d7da !important;
        border-left: 4px solid #dc3545;
    }

    /* Excel-like cell formatting */
    .excel-table-container .stDataFrame td[data-type="currency"] {
        text-align: right;
        font-weight: 600;
        color: #2e7d32;
    }

    .excel-table-container .stDataFrame td[data-type="number"] {
        text-align: right;
    }

    .excel-table-container .stDataFrame td[data-type="status"] {
        text-align: center;
        font-weight: 600;
        text-transform: uppercase;
        font-size: 11px;
        letter-spacing: 0.5px;
    }

    /* Metric cards enhancement */
    .metric-card {
        background: linear-gradient(135deg, #ffffff 0%, #f8f9fa 100%);
        padding: 1.5rem;
        border-radius: 12px;
        box-shadow: 0 4px 20px rgba(0,0,0,0.1);
        border-left: 4px solid var(--primary-color);
        border: 1px solid #dee2e6;
        transition: transform 0.2s ease;
        min-height: 120px;
    }

    /* Dynamic metric card styling based on content */
    .metric-card.has-data {
        background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
        border: 2px solid #dee2e6;
        box-shadow: 0 6px 25px rgba(0,0,0,0.15);
    }

    .metric-card.no-data {
        background: linear-gradient(135deg, #ffffff 0%, #f8f9fa 100%);
        border: 1px solid #e9ecef;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
        min-height: 80px;
        padding: 1rem;
    }

    /* Info boxes */
    .info-box {
        background: linear-gradient(135deg, #e3f2fd 0%, #bbdefb 100%);
        padding: 1rem;
        border-radius: 8px;
        border-left: 4px solid var(--info-color);
        margin: 1rem 0;
        border: 1px solid #90caf9;
    }

    /* Alerts */
    .alert {
        padding: 1rem 1.5rem;
        border-radius: 8px;
        margin: 1rem 0;
        border-left: 4px solid;
    }

    .alert-success {
        background-color: #d4edda;
        border-color: var(--success-color);
        color: #155724;
    }

    .alert-warning {
        background-color: #fff3cd;
        border-color: var(--warning-color);
        color: #856404;
    }

    .alert-info {
        background-color: #d1ecf1;
        border-color: var(--info-color);
        color: #0c5460;
    }

    /* Charts */
    .chart-container {
        background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
        padding: 1.5rem;
        border-radius: 12px;
        box-shadow: 0 4px 20px rgba(0,0,0,0.1);
        margin-bottom: 2rem;
        border: 1px solid #dee2e6;
    }

    /* Navigation */
    .nav-item {
        padding: 0.75rem 1rem;
        margin: 0.25rem 0;
        border-radius: 8px;
        transition: all 0.2s ease;
        cursor: pointer;
    }

    .nav-item:hover {
        background-color: rgba(31, 119, 180, 0.1);
    }

    /* Responsive Design */
    @media (max-width: 768px) {
        .main-header h1 {
            font-size: 2rem;
        }
        .metric-value {
            font-size: 1.5rem;
        }
    }
    </style>
""", unsafe_allow_html=True)

# Main Header
st.markdown("""
<div class="main-header">
    <h1>🚀 Comprehensive Project Management Dashboard</h1>
    <p>Streamline your project lifecycle with intelligent tracking and management</p>
</div>
""", unsafe_allow_html=True)

# Enhanced Sidebar Navigation
st.sidebar.markdown("""
<div style="text-align: center; padding: 1rem 0; border-bottom: 2px solid #e9ecef; margin-bottom: 1rem;">
    <h2 style="color: #2c3e50; margin: 0;">📋 Navigation</h2>
</div>
""", unsafe_allow_html=True)

# Navigation options with descriptions
nav_options = {
    "🏠 Dashboard Overview": "View project statistics and analytics",
    "📝 Submit Project": "Create and submit new projects",
    "🔍 Track Project": "Monitor project status and progress",
    "👨‍💼 Admin Panel": "Review and manage projects",
    "👑 Super User Dashboard": "Advanced analytics and cumulative calculations",
    "📊 Analytics": "Advanced project analytics",
    "📁 File Upload": "Upload and process project files",
    "📦 Material Entry": "Manage project materials and BOQ"
}

page = st.sidebar.selectbox(
    "Choose a page:",
    list(nav_options.keys())
)

# Show description for selected page
st.sidebar.markdown(f"""
<div style="background: #f8f9fa; padding: 1rem; border-radius: 8px; margin-top: 1rem; border-left: 4px solid #17a2b8;">
    <small style="color: #6c757d;">{nav_options[page]}</small>
</div>
""", unsafe_allow_html=True)

# Add some sidebar info
st.sidebar.markdown("---")
st.sidebar.markdown("""
<div style="text-align: center; padding: 1rem;">
    <small style="color: #6c757d;">
        <strong>Project Management System</strong><br>
        Version 2.0<br>
        Streamlit Dashboard
    </small>
</div>
""", unsafe_allow_html=True)

# Dashboard Overview
if page == "🏠 Dashboard Overview":
    # Page Header with Breadcrumb
    st.markdown("""
    <div style="margin-bottom: 2rem;">
        <nav style="margin-bottom: 1rem;">
            <span style="color: #6c757d;">🏠 Dashboard</span> 
            <span style="color: #6c757d;">›</span> 
            <span style="color: #2c3e50; font-weight: 600;">Overview</span>
        </nav>
        <h1 style="color: #2c3e50; margin: 0;">📊 Dashboard Overview</h1>
        <p style="color: #6c757d; margin: 0.5rem 0 0 0;">Real-time project insights and performance metrics</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Get projects data
    projects_df = get_projects()
    
    if not projects_df.empty:
        # Enhanced Statistics Cards
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            total_projects = len(projects_df)
            card_class = "has-data" if total_projects > 0 else "no-data"
            st.markdown(f"""
            <div class="metric-card {card_class}">
                <div class="metric-value">{total_projects}</div>
                <div class="metric-label">📊 Total Projects</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            pending_projects = len(projects_df[projects_df['status'] == 'pending'])
            card_class = "has-data" if pending_projects > 0 else "no-data"
            st.markdown(f"""
            <div class="metric-card {card_class}">
                <div class="metric-value" style="color: #ffc107;">{pending_projects}</div>
                <div class="metric-label">⏳ Pending Review</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            approved_projects = len(projects_df[projects_df['status'] == 'approved'])
            card_class = "has-data" if approved_projects > 0 else "no-data"
            st.markdown(f"""
            <div class="metric-card {card_class}">
                <div class="metric-value" style="color: #28a745;">{approved_projects}</div>
                <div class="metric-label">✅ Approved</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col4:
            total_budget = projects_df['total_cost'].sum()
            card_class = "has-data" if total_budget > 0 else "no-data"
            st.markdown(f"""
            <div class="metric-card {card_class}">
                <div class="metric-value" style="color: #17a2b8;">₹{total_budget:,.0f}</div>
                <div class="metric-label">💰 Total Budget</div>
            </div>
            """, unsafe_allow_html=True)
        
        # Additional Metrics Row
        st.markdown("<br>", unsafe_allow_html=True)
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            rejected_projects = len(projects_df[projects_df['status'] == 'rejected'])
            card_class = "has-data" if rejected_projects > 0 else "no-data"
            st.markdown(f"""
            <div class="metric-card {card_class}">
                <div class="metric-value" style="color: #dc3545;">{rejected_projects}</div>
                <div class="metric-label">❌ Rejected</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            under_review = len(projects_df[projects_df['status'] == 'under_review'])
            card_class = "has-data" if under_review > 0 else "no-data"
            st.markdown(f"""
            <div class="metric-card {card_class}">
                <div class="metric-value" style="color: #6f42c1;">{under_review}</div>
                <div class="metric-label">🔍 Under Review</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            avg_cost = projects_df['total_cost'].mean()
            card_class = "has-data" if avg_cost > 0 else "no-data"
            st.markdown(f"""
            <div class="metric-card {card_class}">
                <div class="metric-value" style="color: #fd7e14;">₹{avg_cost:,.0f}</div>
                <div class="metric-label">📈 Avg Project Cost</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col4:
            approval_rate = (approved_projects / total_projects * 100) if total_projects > 0 else 0
            card_class = "has-data" if approval_rate > 0 else "no-data"
            st.markdown(f"""
            <div class="metric-card {card_class}">
                <div class="metric-value" style="color: #20c997;">{approval_rate:.1f}%</div>
                <div class="metric-label">📊 Approval Rate</div>
            </div>
            """, unsafe_allow_html=True)
        
        # Enhanced Charts Section
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("### 📈 Project Analytics")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.markdown('<div class="chart-container">', unsafe_allow_html=True)
            # Status distribution with enhanced colors
            status_counts = projects_df['status'].value_counts()
            status_colors = {
                'pending': '#ffc107',
                'approved': '#28a745', 
                'rejected': '#dc3545',
                'under_review': '#6f42c1'
            }
            
            fig_status = px.pie(
                values=status_counts.values, 
                names=status_counts.index,
                title="📊 Project Status Distribution",
                color_discrete_map=status_colors,
                hole=0.4
            )
            fig_status.update_layout(
                title_font_size=18,
                legend_title_font_size=14,
                legend_font_size=12,
                font=dict(size=12),
                showlegend=True,
                legend=dict(
                    orientation="v",
                    yanchor="middle",
                    y=0.5,
                    xanchor="left",
                    x=1.01
                )
            )
            st.plotly_chart(fig_status, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)
        
        with col2:
            st.markdown('<div class="chart-container">', unsafe_allow_html=True)
            # Domain distribution with enhanced styling
            domain_counts = projects_df['domain'].value_counts()
            fig_domain = px.bar(
                x=domain_counts.index, 
                y=domain_counts.values,
                title="🏗️ Projects by Domain",
                color=domain_counts.values,
                color_continuous_scale='Blues',
                text=domain_counts.values
            )
            fig_domain.update_traces(
                texttemplate='%{text}',
                textposition='outside',
                marker_line_color='white',
                marker_line_width=1
            )
            fig_domain.update_layout(
                title_font_size=18,
                xaxis_title="Domain",
                yaxis_title="Number of Projects",
                font=dict(size=12),
                xaxis_tickangle=-45
            )
            st.plotly_chart(fig_domain, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)
        
        # Budget by category analysis
        st.markdown('<div class="chart-container">', unsafe_allow_html=True)
        st.subheader("💰 Budget by Category")
        
        # Get all materials data to analyze budget by category
        conn = sqlite3.connect('project_management.db')
        materials_df = pd.read_sql_query("SELECT * FROM project_materials", conn)
        conn.close()
        
        if not materials_df.empty:
            # Group by category and sum the amounts
            category_budget = materials_df.groupby('category')['amount_inr'].sum().reset_index()
            category_budget = category_budget.sort_values('amount_inr', ascending=False)
            
            fig_category = px.bar(
                category_budget,
                x='category',
                y='amount_inr',
                title="Budget Distribution by Category",
                color='amount_inr',
                color_continuous_scale='Blues',
                labels={'amount_inr': 'Amount (₹)', 'category': 'Category'}
            )
            fig_category.update_layout(
                title_font_size=20,
                xaxis_title="Category",
                yaxis_title="Budget (₹)",
                xaxis_tickangle=-45
            )
            st.plotly_chart(fig_category, use_container_width=True)
        else:
            st.info("No category budget data available yet.")
        st.markdown('</div>', unsafe_allow_html=True)
        
        # Recent projects table with enhanced styling
        st.markdown('<div class="dashboard-card">', unsafe_allow_html=True)
        st.subheader("📋 Recent Projects")
        display_df = projects_df[['tracking_id', 'project_name', 'domain', 'priority', 'total_cost', 'status', 'submitted_at']].head(10)
        
        # Format the dataframe for better display
        display_df['total_cost'] = display_df['total_cost'].apply(lambda x: f"₹{x:,.2f}")
        display_df['status'] = display_df['status'].apply(lambda x: x.title())
        display_df['submitted_at'] = pd.to_datetime(display_df['submitted_at']).dt.strftime('%Y-%m-%d %H:%M')
        
        # Rename columns for better display
        display_df.columns = ['Tracking ID', 'Project Name', 'Domain', 'Priority', 'Total Cost', 'Status', 'Submitted At']
        
        st.dataframe(display_df, use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)
    
    else:
        st.info("No projects found. Submit your first project to get started!")

# Submit Project
elif page == "📝 Submit Project":
    # Page Header with Breadcrumb
    st.markdown("""
    <div style="margin-bottom: 2rem;">
        <nav style="margin-bottom: 1rem;">
            <span style="color: #6c757d;">🏠 Dashboard</span> 
            <span style="color: #6c757d;">›</span> 
            <span style="color: #2c3e50; font-weight: 600;">Submit Project</span>
        </nav>
        <h1 style="color: #2c3e50; margin: 0;">📝 Submit New Project</h1>
        <p style="color: #6c757d; margin: 0.5rem 0 0 0;">Create and submit a new project for review and approval</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Step 1: Project Basic Details
    if "project_info_submitted" not in st.session_state:
        st.session_state["project_info_submitted"] = False
    if "project_basic" not in st.session_state:
        st.session_state["project_basic"] = {}
    if "project_materials_data" not in st.session_state:
        st.session_state["project_materials_data"] = []

    if not st.session_state["project_info_submitted"]:
        st.markdown('<div class="form-container">', unsafe_allow_html=True)
        st.markdown("### 📋 Step 1: Project Basic Information")
        st.markdown('<div class="form-section">', unsafe_allow_html=True)
        
        with st.form("project_form_step1"):
            col1, col2 = st.columns(2)
            with col1:
                st.markdown('<div class="topic-section">', unsafe_allow_html=True)
                st.markdown("#### 🏗️ Project Details")
                project_name = st.text_input("Project Name *", placeholder="Enter project name")
                domain = st.selectbox("Domain *", ["Civil Engineering", "Mechanical Engineering", "Electrical Engineering", "Multi-Domain"])
                priority = st.selectbox("Priority *", ["Low", "Medium", "High", "Urgent"])
                project_description = st.text_area("Project Description *", height=100)
                st.markdown('</div>', unsafe_allow_html=True)
                
                st.markdown('<div class="topic-section">', unsafe_allow_html=True)
                st.markdown("#### 📅 Timeline")
                start_date = st.date_input("Start Date *", value=date.today())
                end_date = st.date_input("End Date *", value=date.today())
                st.markdown('</div>', unsafe_allow_html=True)
                
            with col2:
                st.markdown('<div class="topic-section">', unsafe_allow_html=True)
                st.markdown("#### 👤 Contact Information")
                department = st.text_input("Department", placeholder="Your department")
                contact_email = st.text_input("Contact Email", placeholder="your.email@company.com")
                contact_phone = st.text_input("Contact Phone", placeholder="+91 9876543210")
            submitted_by = st.text_input("Submitted By *", placeholder="Your name")
            st.markdown('</div>', unsafe_allow_html=True)
                
            st.markdown('<div class="topic-section">', unsafe_allow_html=True)
            st.markdown("#### 📄 Justification")
            justification_type = st.radio("Justification Type:", ["Text Input", "File Upload"], horizontal=True)
                
            if justification_type == "Text Input":
                    justification = st.text_area("Justification *", height=100, placeholder="Provide justification for this project...")
            else:
                    uploaded_file = st.file_uploader("Upload Justification Document *", type=['pdf', 'docx', 'txt'], help="Upload PDF, DOCX, or TXT file")
                    justification = uploaded_file.name if uploaded_file else ""
            st.markdown('</div>', unsafe_allow_html=True)
            
            col1, col2, col3 = st.columns([1, 1, 1])
            with col2:
                go_next = st.form_submit_button("🚀 Proceed to BOM/Material Entry →", use_container_width=True)
            
            if go_next:
                required = [project_name, project_description, submitted_by]
                if justification_type == "Text Input":
                    required.append(justification)
                else:
                    required.append(uploaded_file)
                
                if all(required):
                    # Handle file upload if applicable
                    justification_content = justification
                    if justification_type == "File Upload" and uploaded_file:
                        # Save uploaded file
                        upload_dir = "data/uploads/justifications"
                        os.makedirs(upload_dir, exist_ok=True)
                        file_path = os.path.join(upload_dir, uploaded_file.name)
                        with open(file_path, "wb") as f:
                            f.write(uploaded_file.getbuffer())
                        justification_content = f"File uploaded: {uploaded_file.name}"
                    
                    st.session_state["project_basic"] = {
                        'project_name': project_name,
                        'project_description': project_description,
                        'domain': domain,
                        'priority': priority,
                        'start_date': start_date,
                        'end_date': end_date,
                        'department': department,
                        'contact_email': contact_email,
                        'contact_phone': contact_phone,
                        'justification': justification_content,
                        'justification_type': justification_type,
                        'submitted_by': submitted_by
                    }
                    st.session_state["project_info_submitted"] = True
                    st.success("✅ Basic project info saved. Now add BOM/Material details.")
                else:
                    st.error("❌ Please fill all required fields.")
        
        st.markdown('</div>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    # Step 2: BOM/Materials Entry
    if st.session_state["project_info_submitted"]:
        st.markdown('<div class="form-container">', unsafe_allow_html=True)
        st.markdown("### 📦 Step 2: BOM/Material Items Entry")
        st.markdown('<div class="form-section">', unsafe_allow_html=True)
        
        # Define the categories/topics with their subtopics
        topics_with_subtopics = {
            "1. Supply": [
                "Fabrication (without machining)",
                "Fabrication (with machining)",
                "Raw Materials",
                "Components & Parts",
                "Equipment Supply",
                "Tools & Instruments"
            ],
            "2. Erection": [
                "Structural Erection",
                "Mechanical Erection",
                "Electrical Erection",
                "Piping Erection",
                "Instrumentation Erection",
                "Civil Works"
            ],
            "3. Erection & Commissioning": [
                "Installation & Testing",
                "Commissioning Services",
                "Start-up Support",
                "Performance Testing",
                "Training & Documentation",
                "Warranty Support"
            ],
            "4. Running Expenses": [
                "Fuel & Energy",
                "Maintenance Materials",
                "Spare Parts",
                "Consumables",
                "Utilities",
                "Operational Supplies"
            ],
            "5. Project Management Service PMC": [
                "Project Planning",
                "Quality Control",
                "Safety Management",
                "Progress Monitoring",
                "Coordination Services",
                "Documentation Management"
            ],
            "6. Travels & Others": [
                "Travel Expenses",
                "Accommodation",
                "Transportation",
                "Communication",
                "Miscellaneous Expenses",
                "Contingency"
            ],
            "7. Bonds & Guarantee": [
                "Performance Bond",
                "Bank Guarantee",
                "Warranty Bond",
                "Insurance",
                "Retention Money",
                "Security Deposits"
            ],
            "8. Statutory expenses": [
                "Taxes & Duties",
                "Licenses & Permits",
                "Regulatory Compliance",
                "Environmental Clearances",
                "Safety Certifications",
                "Legal Fees"
            ],
            "9. Over head - Chennai office": [
                "Office Rent",
                "Utilities",
                "Staff Salaries",
                "Administrative Costs",
                "IT Infrastructure",
                "General Expenses"
            ],
            "10. Export": [
                "Export Documentation",
                "Shipping & Logistics",
                "Customs Clearance",
                "International Compliance",
                "Currency Exchange",
                "Export Incentives"
            ],
            "11. Engineering & PMC Support - Third Party": [
                "Design Engineering",
                "Technical Consultancy",
                "Third Party Inspection",
                "Testing Services",
                "Certification Services",
                "Expert Consultation"
            ],
            "12. Service Charges": [
                "Professional Services",
                "Consultancy Fees",
                "Technical Support",
                "Maintenance Services",
                "Training Services",
                "After Sales Support"
            ],
            "13. Royaltee": [
                "Technology License",
                "Patent Fees",
                "Intellectual Property",
                "Software Licenses",
                "Brand Licensing",
                "Technical Know-how"
            ],
            "14. Contigencies": [
                "Project Contingency",
                "Price Escalation",
                "Scope Changes",
                "Risk Mitigation",
                "Unforeseen Events",
                "Buffer Amount"
            ],
            "15. Net Margin": [
                "Profit Margin",
                "Overhead Recovery",
                "Risk Premium",
                "Company Profit",
                "Return on Investment",
                "Financial Returns"
            ]
        }
        
        categories = list(topics_with_subtopics.keys())
        units_options = ["kgs", "MT", "mtrs", "Sq.ft", "Sq.Mtr", "Litres", "RMT", "No,s"]
        nos_options = list(range(1, 21))
        source_type_options = ["Vendor Quote", "Company Costing", "Free Issue"]
        payment_schedule_options = ["Ontime", "Monthly"]

        with st.form("material_entry_proj_form_step2"):
            # First select the category/topic
            category = st.selectbox("Select Category/Topic", categories, key="category_proj2")
            
            # Then select subtopic based on the selected category
            subtopics = topics_with_subtopics[category]
            subtopic = st.selectbox("Select Sub-topic", subtopics, key="subtopic_proj2")
            
            # Then enter additional description details
            description = st.text_input("Additional Description (Optional)", placeholder="Add more details if needed", key="desc_proj2")
            
            mcol1, mcol2, mcol3 = st.columns([3, 1, 1])
            with mcol1:
                units = st.selectbox("Units/Qty", units_options, key='units_proj2')
            with mcol2:
                nos = st.selectbox("Nos", nos_options, key='nos_proj2')
            with mcol3:
                unit_price = st.number_input("Unit Price Total Amount", min_value=0.0, step=0.01, format="%.2f", key='unit_price_proj2')
            row2_col1, row2_col2, row2_col3 = st.columns([2,2,2])
            with row2_col1:
                source_type = st.selectbox("Source/Type", source_type_options, key='stype_proj2')
            with row2_col2:
                payment_schedule = st.selectbox("Payment Schedule", payment_schedule_options, key='pay_sched_proj2')
            with row2_col3:
                st.empty()  # Empty column for alignment
            amount_inr = nos * unit_price
            st.markdown(f"##### Amount INR: <span style='color: #3973ac; font-size: 22px;'>₹{amount_inr:,.2f}</span>", unsafe_allow_html=True)
            
            # Justification section for project materials
            st.markdown("#### 📄 Justification for this Material Entry")
            proj_justification_type = st.radio("Justification Type:", ["Text Input", "File Upload"], horizontal=True, key="proj_mat_justification_type")
            
            if proj_justification_type == "Text Input":
                proj_material_justification = st.text_area("Justification *", height=80, placeholder="Provide justification for this material entry...", key="proj_mat_justification_text")
            else:
                proj_material_justification_file = st.file_uploader("Upload Justification Document *", type=['pdf', 'docx', 'txt'], help="Upload PDF, DOCX, or TXT file", key="proj_mat_justification_file")
                proj_material_justification = proj_material_justification_file.name if proj_material_justification_file else ""
            
            add_mat = st.form_submit_button("Add Material to Project")
        if add_mat:
            # Combine subtopic and additional description
            full_description = f"{subtopic}"
            if description:
                full_description += f" - {description}"
            
            # Handle justification
            justification_content = proj_material_justification
            justification_file_path = ""
            
            if proj_justification_type == "File Upload" and proj_material_justification_file:
                # Save uploaded file
                upload_dir = "data/uploads/project_material_justifications"
                os.makedirs(upload_dir, exist_ok=True)
                file_path = os.path.join(upload_dir, proj_material_justification_file.name)
                with open(file_path, "wb") as f:
                    f.write(proj_material_justification_file.getbuffer())
                justification_file_path = file_path
                justification_content = f"File uploaded: {proj_material_justification_file.name}"
            
            required = [subtopic, units, nos, source_type, payment_schedule]
            if proj_justification_type == "Text Input":
                required.append(proj_material_justification)
            else:
                required.append(proj_material_justification_file)
            
            if all(required) and unit_price > 0:
                entry = {
                    "Category": category,
                    "Sub-topic": subtopic,
                    "Description": full_description,
                    "Units/Qty": units,
                    "Nos": nos,
                    "Source/Type": source_type,
                    "Payment Schedule": payment_schedule,
                    "Unit Price Total Amount": unit_price,
                    "Amount INR": amount_inr,
                    "Justification": justification_content,
                    "Justification Type": proj_justification_type,
                    "Justification File Path": justification_file_path
                }
                st.session_state["project_materials_data"].append(entry)
                st.success(f"Material entry added to project: {category} - {subtopic}!")
            else:
                st.error("❌ Please fill all required fields and enter Unit Price > 0.")
        # Show review table with categories and row numbers
        if st.session_state["project_materials_data"]:
            st.markdown("#### Added Material Details (Please Review)")
            
            # Group by category for display
            categories_in_data = sorted(set(item["Category"] for item in st.session_state["project_materials_data"]))
            
            for cat in categories_in_data:
                with st.expander(f"{cat}", expanded=True):
                    cat_data = [item for item in st.session_state["project_materials_data"] if item["Category"] == cat]
                    cat_df = pd.DataFrame(cat_data)
                    
                    # Add row numbers
                    cat_df.index = cat_df.index + 1
                    cat_df.reset_index(inplace=True)
                    cat_df.rename(columns={"index": "S.No."}, inplace=True)
                    
                    # Calculate totals for this category
                    total_amount = sum(item["Amount INR"] for item in cat_data)
                    
                    # Display the data table with subtopic information
                    display_columns = ["S.No.", "Sub-topic", "Description", "Units/Qty", "Nos", "Unit Price Total Amount", "Amount INR"]
                    st.dataframe(cat_df[display_columns], use_container_width=True)
                    
                    st.markdown(f"**Total for {cat}: ₹{total_amount:,.2f}**")
                    
                    # Add delete buttons for entries in this category
                    if st.button(f"Delete All in {cat}", key=f"del_proj_cat_{cat}"):
                        st.session_state["project_materials_data"] = [
                            item for item in st.session_state["project_materials_data"] if item["Category"] != cat
                        ]
                        st.success(f"All entries in {cat} deleted!")
                        st.experimental_rerun()
            
            # Show grand total
            grand_total = sum(item["Amount INR"] for item in st.session_state["project_materials_data"])
            st.markdown(f"### Grand Total: ₹{grand_total:,.2f}")
        submitted = st.button("Submit Full Project (with BOM)")
        if submitted:
            if not st.session_state["project_materials_data"]:
                st.error("At least one material/BOM entry is required.")
            else:
                # Save both project and materials
                project_data = st.session_state["project_basic"]
                tracking_id = save_project(project_data)
                # Now get the new project's internal id
                conn = sqlite3.connect('project_management.db')
                cursor = conn.cursor()
                cursor.execute("SELECT id FROM projects WHERE tracking_id = ?", (tracking_id,))
                projidrow = cursor.fetchone()
                conn.close()
                if projidrow:
                    new_proj_id = projidrow[0]
                    save_materials(new_proj_id, st.session_state["project_materials_data"])
                    st.success(f"✅ Project and Materials submitted successfully!")
                    # Notify admin for review
                    add_notification(new_proj_id, "project_submitted", f"Project Submitted: {tracking_id}", f"Project '{project_data['project_name']}' submitted with {len(st.session_state['project_materials_data'])} BOQ rows.")
                else:
                    st.error("❌ Project saved, but could not store materials (Internal error)")
                
                # Show tracking ID and add notification regardless of materials save success
                st.success(f"📋 Tracking ID: **{tracking_id}**")
                st.info("💡 Use this tracking ID to monitor your project status.")
                # User-side notification already added above when proj id known
                st.session_state["project_info_submitted"] = False
                st.session_state["project_materials_data"] = []
                st.session_state["project_basic"] = {}

# Track Project
elif page == "🔍 Track Project":
    # Page Header with Breadcrumb
    st.markdown("""
    <div style="margin-bottom: 2rem;">
        <nav style="margin-bottom: 1rem;">
            <span style="color: #6c757d;">🏠 Dashboard</span> 
            <span style="color: #6c757d;">›</span> 
            <span style="color: #2c3e50; font-weight: 600;">Track Project</span>
        </nav>
        <h1 style="color: #2c3e50; margin: 0;">🔍 Track Project Status</h1>
        <p style="color: #6c757d; margin: 0.5rem 0 0 0;">Monitor your project progress and status updates</p>
    </div>
    """, unsafe_allow_html=True)
    
    tracking_id = st.text_input("Enter Tracking ID", placeholder="e.g., PRJ251019A1B2")
    
    if tracking_id:
        project = get_project_by_tracking_id(tracking_id)
        
        if project:
            st.markdown(f"""
            <div class="alert alert-success">
                <strong>✅ Project Found:</strong> {project[2]}
            </div>
            """, unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown('<div class="form-container">', unsafe_allow_html=True)
                st.markdown("### 📋 Project Information")
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #e3f2fd 0%, #bbdefb 100%); padding: 1.5rem; border-radius: 8px; margin: 1rem 0; border: 1px solid #90caf9; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
                    <p style="color: #1565c0; margin: 0.5rem 0;"><strong>Tracking ID:</strong> <code style="background: rgba(255,255,255,0.8); padding: 0.2rem 0.5rem; border-radius: 4px;">{project[1]}</code></p>
                    <p style="color: #2c3e50; margin: 0.5rem 0;"><strong>Project Name:</strong> {project[2]}</p>
                    <p style="color: #2c3e50; margin: 0.5rem 0;"><strong>Domain:</strong> {project[4]}</p>
                    <p style="color: #2c3e50; margin: 0.5rem 0;"><strong>Priority:</strong> {project[5]}</p>
                    <p style="color: #2c3e50; margin: 0.5rem 0;"><strong>Total Cost:</strong> ₹{project[12]:,.2f}</p>
                    <p style="color: #2c3e50; margin: 0.5rem 0;"><strong>Submitted By:</strong> {project[22]}</p>
                    <p style="color: #2c3e50; margin: 0.5rem 0;"><strong>Submitted On:</strong> {project[23]}</p>
                </div>
                """, unsafe_allow_html=True)
                st.markdown('</div>', unsafe_allow_html=True)
            
            with col2:
                st.markdown('<div class="form-container">', unsafe_allow_html=True)
                st.markdown("### 📊 Current Status")
                status = project[21]
                
                # Status badge
                status_badge_class = f"status-{status.replace('_', '-')}"
                st.markdown(f"""
                <div style="margin: 1rem 0;">
                    <span class="status-badge {status_badge_class}">{status.replace('_', ' ').title()}</span>
                </div>
                """, unsafe_allow_html=True)
                
                # Additional info
                if len(project) > 25 and project[25]:  # reviewed_by
                    st.markdown(f"""
                    <div style="background: linear-gradient(135deg, #e8f5e8 0%, #c8e6c9 100%); padding: 1.5rem; border-radius: 8px; margin: 1rem 0; border: 1px solid #81c784; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
                        <p style="color: #2e7d32; margin: 0.5rem 0;"><strong>Reviewed By:</strong> {project[25]}</p>
                        <p style="color: #2e7d32; margin: 0.5rem 0;"><strong>Review Date:</strong> {project[26]}</p>
                    </div>
                    """, unsafe_allow_html=True)
                
                if len(project) > 27 and project[27]:  # review_comments
                    st.markdown(f"""
                    <div style="background: linear-gradient(135deg, #fff3e0 0%, #ffcc02 100%); padding: 1rem; border-radius: 8px; margin: 1rem 0; border: 1px solid #ffb74d; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
                        <strong style="color: #e65100;">Review Comments:</strong><br>
                        <em style="color: #bf360c;">{project[27]}</em>
                    </div>
                    """, unsafe_allow_html=True)
                st.markdown('</div>', unsafe_allow_html=True)
            
            # Timeline
            st.subheader("📅 Project Timeline")
            timeline_data = [
                {"Event": "Project Submitted", "Date": project[23], "Status": "Completed"},
                {"Event": "Under Review", "Date": project[26] if project[26] else "Pending", "Status": "In Progress" if status == 'pending' else "Completed"},
                {"Event": "Decision Made", "Date": project[26] if project[26] else "Pending", "Status": "Completed" if status in ['approved', 'rejected'] else "Pending"}
            ]
            
            timeline_df = pd.DataFrame(timeline_data)
            st.dataframe(timeline_df, use_container_width=True)
        
        else:
            st.error("❌ Project not found. Please check your tracking ID.")

# Admin Panel
elif page == "👨‍💼 Admin Panel":
    st.markdown("""
    <div style="margin-bottom: 2rem;">
        <nav style="margin-bottom: 1rem;">
            <span style="color: #6c757d;">🏠 Dashboard</span> 
            <span style="color: #6c757d;">›</span> 
            <span style="color: #2c3e50; font-weight: 600;">Admin Panel (Detailed Review)</span>
        </nav>
        <h1 style="color: #2c3e50; margin: 0;">👨‍💼 Admin Panel</h1>
        <p style="color: #6c757d; margin: 0.5rem 0 0 0;">Comprehensive BOQ/Materials review and project validation</p>
    </div>
    """, unsafe_allow_html=True)
    
    admin_password = st.text_input("Admin Password", type="password", placeholder="Enter admin password")
    
    if admin_password == "admin123":
        st.success("✅ Admin access granted!")
        projects_df = get_projects()
        
        if not projects_df.empty:
            # Admin Dashboard Overview
            st.markdown('<div class="form-container">', unsafe_allow_html=True)
            st.markdown("### 📊 Admin Dashboard Overview")
            
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                pending_count = len(projects_df[projects_df['status'] == 'pending'])
                card_class = "has-data" if pending_count > 0 else "no-data"
                st.markdown(f"""
                <div class="metric-card {card_class}" style="border-left-color: #ffc107;">
                    <div class="metric-value" style="color: #ffc107;">{pending_count}</div>
                    <div class="metric-label">⏳ Pending Review</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                under_review_count = len(projects_df[projects_df['status'] == 'under_review'])
                card_class = "has-data" if under_review_count > 0 else "no-data"
                st.markdown(f"""
                <div class="metric-card {card_class}" style="border-left-color: #6f42c1;">
                    <div class="metric-value" style="color: #6f42c1;">{under_review_count}</div>
                    <div class="metric-label">🔍 Under Review</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                approved_count = len(projects_df[projects_df['status'] == 'approved'])
                card_class = "has-data" if approved_count > 0 else "no-data"
                st.markdown(f"""
                <div class="metric-card {card_class}" style="border-left-color: #28a745;">
                    <div class="metric-value" style="color: #28a745;">{approved_count}</div>
                    <div class="metric-label">✅ Approved</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col4:
                rejected_count = len(projects_df[projects_df['status'] == 'rejected'])
                card_class = "has-data" if rejected_count > 0 else "no-data"
                st.markdown(f"""
                <div class="metric-card {card_class}" style="border-left-color: #dc3545;">
                    <div class="metric-value" style="color: #dc3545;">{rejected_count}</div>
                    <div class="metric-label">❌ Rejected</div>
                </div>
                """, unsafe_allow_html=True)
            
            st.markdown('</div>', unsafe_allow_html=True)
            
            # Project Selection for BOQ Review
            st.markdown('<div class="form-container">', unsafe_allow_html=True)
            st.markdown("### 🔍 Select Project for BOQ Review")
            
            # Create a clean dropdown for project selection
            project_options = []
            for idx, project in projects_df.iterrows():
                status_emoji = {
                    'pending': '⏳',
                    'under_review': '🔍', 
                    'approved': '✅',
                    'rejected': '❌'
                }.get(project['status'], '📋')
                
                project_options.append(f"{status_emoji} [{project['tracking_id']}] {project['project_name']} - {project['status'].title()}")
            
            if project_options:
                selected_project_idx = st.selectbox(
                    "Choose a project to review:",
                    range(len(project_options)),
                    format_func=lambda x: project_options[x],
                    key="project_selector"
                )
            else:
                st.warning("No projects available for review.")
                selected_project_idx = None
            
            if selected_project_idx is not None:
                selected_project = projects_df.iloc[selected_project_idx]
                
                # Project Basic Information
                st.markdown("---")
                st.markdown("### 📋 Project Basic Information")
                
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown(f"""
                    <div style="background: linear-gradient(135deg, #e3f2fd 0%, #bbdefb 100%); padding: 1.5rem; border-radius: 8px; margin: 1rem 0; border: 1px solid #90caf9; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
                        <p style="color: #1565c0; margin: 0.5rem 0;"><strong>Tracking ID:</strong> <code style="background: rgba(255,255,255,0.8); padding: 0.2rem 0.5rem; border-radius: 4px;">{selected_project['tracking_id']}</code></p>
                        <p style="color: #2c3e50; margin: 0.5rem 0;"><strong>Project Name:</strong> {selected_project['project_name']}</p>
                        <p style="color: #2c3e50; margin: 0.5rem 0;"><strong>Domain:</strong> {selected_project['domain']}</p>
                        <p style="color: #2c3e50; margin: 0.5rem 0;"><strong>Priority:</strong> {selected_project['priority']}</p>
                        <p style="color: #2c3e50; margin: 0.5rem 0;"><strong>Total Cost:</strong> ₹{selected_project['total_cost']:,.2f}</p>
                    </div>
                    """, unsafe_allow_html=True)
                
                with col2:
                    st.markdown(f"""
                    <div style="background: linear-gradient(135deg, #e8f5e8 0%, #c8e6c9 100%); padding: 1.5rem; border-radius: 8px; margin: 1rem 0; border: 1px solid #81c784; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
                        <p style="color: #2e7d32; margin: 0.5rem 0;"><strong>Submitted By:</strong> {selected_project['submitted_by']}</p>
                        <p style="color: #2e7d32; margin: 0.5rem 0;"><strong>Department:</strong> {selected_project.get('department', 'N/A')}</p>
                        <p style="color: #2e7d32; margin: 0.5rem 0;"><strong>Contact:</strong> {selected_project.get('contact_email', 'N/A')}</p>
                        <p style="color: #2e7d32; margin: 0.5rem 0;"><strong>Submitted On:</strong> {selected_project['submitted_at']}</p>
                    </div>
                    """, unsafe_allow_html=True)
                
                # Project Description and Justification
                st.markdown("### 📄 Project Details")
                st.markdown(f"""
                <div style="background: linear-gradient(135deg, #fff3e0 0%, #ffcc02 100%); padding: 1.5rem; border-radius: 8px; margin: 1rem 0; border: 1px solid #ffb74d; box-shadow: 0 2px 10px rgba(0,0,0,0.1);">
                    <p style="color: #e65100; margin: 0.5rem 0;"><strong>Description:</strong></p>
                    <p style="color: #bf360c; margin: 0.5rem 0;">{selected_project['project_description']}</p>
                    <p style="color: #e65100; margin: 0.5rem 0;"><strong>Justification:</strong></p>
                    <p style="color: #bf360c; margin: 0.5rem 0;">{selected_project.get('justification', 'No justification provided')}</p>
                </div>
                """, unsafe_allow_html=True)
                
                # BOQ/Materials Review Section
                st.markdown("---")
                st.markdown("### 📦 BOQ/Materials Data Submitted by User")
                
                materials_df = get_materials_by_project(selected_project['id'])
                
                if not materials_df.empty:
                    st.markdown(f"**Project:** {selected_project['project_name']} | **Tracking ID:** {selected_project['tracking_id']}")
                    st.markdown(f"**Total Material Entries:** {len(materials_df)}")
                    
                    # Create clean BOQ table for admin review
                    boq_display_df = materials_df.copy()
                    
                    # Add row numbers
                    boq_display_df['S.No.'] = range(1, len(boq_display_df) + 1)
                    
                    # Select and rename columns for admin review
                    column_order = [
                        'S.No.', 'category', 'subtopic', 'description', 'units_qty', 'nos', 
                        'unit_price', 'amount_inr', 'source_type', 'payment_schedule', 
                        'justification', 'status'
                    ]
                    
                    # Select and rename columns
                    boq_display_df = boq_display_df[column_order]
                    boq_display_df.columns = [
                        'S.No.', 'Category', 'Sub-topic', 'Description', 'Units/Qty', 'Nos',
                        'Unit Price (₹)', 'Amount (₹)', 'Source/Type', 'Payment Schedule',
                        'Justification', 'Current Status'
                    ]
                    
                    # Format currency columns
                    boq_display_df['Unit Price (₹)'] = boq_display_df['Unit Price (₹)'].apply(lambda x: f"₹{x:,.2f}")
                    boq_display_df['Amount (₹)'] = boq_display_df['Amount (₹)'].apply(lambda x: f"₹{x:,.2f}")
                    
                    # Display the BOQ table
                    st.markdown("#### 📊 BOQ Data Table")
                    st.dataframe(
                        boq_display_df,
                        use_container_width=True,
                        height=min(500, (len(boq_display_df) + 1) * 35 + 3),
                        hide_index=True
                    )
                    
                    # Individual Row Status Management
                    st.markdown("---")
                    st.markdown("#### ⚖️ Individual Row Status Management")
                    
                    # Create a form for row-level status updates
                    with st.form("row_status_form"):
                        st.markdown("**Set status for each BOQ item:**")
                        
                        # Track if any changes were made
                        changes_made = False
                        
                        for idx, material in materials_df.iterrows():
                            col1, col2, col3, col4 = st.columns([2, 1, 1, 1])
                            
                            with col1:
                                st.markdown(f"**Row {idx+1}:** {material['subtopic']} - ₹{material['amount_inr']:,.2f}")
                            
                            with col2:
                                # Status dropdown for each row
                                current_status = material.get('status', 'pending')
                                new_status = st.selectbox(
                                    "Status",
                                    ['pending', 'under_review', 'approved', 'rejected'],
                                    index=['pending', 'under_review', 'approved', 'rejected'].index(current_status),
                                    key=f"status_{material['id']}"
                                )
                            
                            with col3:
                                # Review comments for each row
                                current_comments = material.get('review_comments', '')
                                review_comments = st.text_area(
                                    "Comments",
                                    value=current_comments,
                                    height=60,
                                    key=f"comments_{material['id']}",
                                    placeholder="Enter review comments..."
                                )
                            
                            with col4:
                                # Finalize checkbox
                                is_finalized = material.get('finalized', 0) == 1
                                finalize = st.checkbox(
                                    "Finalize",
                                    value=is_finalized,
                                    key=f"finalize_{material['id']}"
                                )
                            
                            st.markdown("---")
                        
                        # Submit button for all changes
                        submitted = st.form_submit_button("💾 Update All Row Statuses", use_container_width=True)
                        
                        if submitted:
                            # Update each row's status
                            for idx, material in materials_df.iterrows():
                                material_id = material['id']
                                new_status = st.session_state.get(f"status_{material_id}", material.get('status', 'pending'))
                                review_comments = st.session_state.get(f"comments_{material_id}", '')
                                finalize = st.session_state.get(f"finalize_{material_id}", False)
                                
                                # Only update if there are changes
                                if (new_status != material.get('status', 'pending') or 
                                    review_comments != material.get('review_comments', '') or
                                    finalize != (material.get('finalized', 0) == 1)):
                                    
                                    update_material_status(material_id, new_status, review_comments, 'Admin', finalize)
                                    changes_made = True
                            
                            if changes_made:
                                st.success("✅ All row statuses updated successfully!")
                                st.rerun()
                            else:
                                st.info("ℹ️ No changes detected.")
                    
                    # Show overall project status based on individual rows
                    st.markdown("---")
                    st.markdown("#### 📊 Overall Project Status Summary")
                    
                    # Count statuses
                    status_counts = materials_df['status'].value_counts().to_dict()
                    approved_count = status_counts.get('approved', 0)
                    pending_count = status_counts.get('pending', 0)
                    under_review_count = status_counts.get('under_review', 0)
                    rejected_count = status_counts.get('rejected', 0)
                    total_rows = len(materials_df)
                    
                    # Show status summary
                    col1, col2, col3, col4, col5 = st.columns(5)
                    with col1:
                        st.metric("Total Rows", total_rows)
                    with col2:
                        st.metric("✅ Approved", approved_count)
                    with col3:
                        st.metric("⏳ Pending", pending_count)
                    with col4:
                        st.metric("🔍 Under Review", under_review_count)
                    with col5:
                        st.metric("❌ Rejected", rejected_count)
                    
                    # Determine overall project status
                    if approved_count == total_rows:
                        overall_status = "All Approved"
                        status_color = "green"
                    elif rejected_count > 0:
                        overall_status = "Has Rejections"
                        status_color = "red"
                    elif under_review_count > 0:
                        overall_status = "Under Review"
                        status_color = "orange"
                    else:
                        overall_status = "Pending Review"
                        status_color = "yellow"
                    
                    st.markdown(f"**Overall Project Status:** <span style='color: {status_color}; font-weight: bold;'>{overall_status}</span>", unsafe_allow_html=True)
                    
                    # Admin Action Buttons
                    st.markdown("---")
                    st.markdown("#### ⚖️ Final Project Decision")
                    
                    # Review comments
                    review_comments = st.text_area(
                        "Final Review Comments",
                        placeholder="Enter your final review comments here...",
                        height=100,
                        key=f"final_review_{selected_project['id']}"
                    )
                    
                    # Action buttons based on overall status
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        # Approve button - only enabled if all rows are approved
                        approve_disabled = approved_count != total_rows
                        if st.button("✅ Approve Project", key=f"approve_{selected_project['id']}", use_container_width=True, disabled=approve_disabled):
                            if review_comments.strip():
                                update_project_status(selected_project['tracking_id'], 'approved', review_comments, 'Admin')
                            else:
                                update_project_status(selected_project['tracking_id'], 'approved', 'Approved by admin', 'Admin')
                            st.success("✅ Project approved!")
                            st.rerun()
                        if approve_disabled:
                            st.caption("All rows must be approved first")
                    
                    with col2:
                        # Reject button - always available
                        if st.button("❌ Reject Project", key=f"reject_{selected_project['id']}", use_container_width=True):
                            if review_comments.strip():
                                update_project_status(selected_project['tracking_id'], 'rejected', review_comments, 'Admin')
                                st.error("❌ Project rejected with comments!")
                            else:
                                update_project_status(selected_project['tracking_id'], 'rejected', 'Rejected by admin', 'Admin')
                                st.error("❌ Project rejected!")
                            st.rerun()
                    
                    with col3:
                        # Mark under review - only if there are under_review items
                        review_disabled = under_review_count == 0
                        if st.button("🔍 Mark Under Review", key=f"review_{selected_project['id']}", use_container_width=True, disabled=review_disabled):
                            if review_comments.strip():
                                update_project_status(selected_project['tracking_id'], 'under_review', review_comments, 'Admin')
                            else:
                                update_project_status(selected_project['tracking_id'], 'under_review', 'Marked under review by admin', 'Admin')
                            st.info("🔍 Project marked under review!")
                            st.rerun()
                        if review_disabled:
                            st.caption("No items under review")
                    
                    with col4:
                        # Request more info - only if there are pending items
                        info_disabled = pending_count == 0
                        if st.button("📋 Request More Info", key=f"info_{selected_project['id']}", use_container_width=True, disabled=info_disabled):
                            if review_comments.strip():
                                update_project_status(selected_project['tracking_id'], 'pending', f"More information requested: {review_comments}", 'Admin')
                                st.warning("📋 More information requested!")
                            else:
                                st.warning("Please provide comments when requesting more information.")
                        if info_disabled:
                            st.caption("No pending items")
                    
                    # Export options
                    col_export1, col_export2, col_export3 = st.columns(3)
                    with col_export1:
                        # Export to CSV
                        csv_data = boq_display_df.to_csv(index=False)
                        st.download_button(
                            label="📊 Export to CSV",
                            data=csv_data,
                            file_name=f"BOQ_{selected_project['tracking_id']}.csv",
                            mime="text/csv",
                            use_container_width=True
                        )
                    
                    with col_export2:
                        # Export to Excel (if openpyxl is available)
                        try:
                            import io
                            from openpyxl import Workbook
                            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                            
                            # Create Excel workbook
                            wb = Workbook()
                            ws = wb.active
                            ws.title = "BOQ Data"
                            
                            # Add headers
                            headers = list(boq_display_df.columns)
                            for col_num, header in enumerate(headers, 1):
                                cell = ws.cell(row=1, column=col_num, value=header)
                                cell.font = Font(bold=True, color="FFFFFF")
                                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                                cell.alignment = Alignment(horizontal="center", vertical="center")
                                cell.border = Border(
                                    left=Side(style='thin'),
                                    right=Side(style='thin'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin')
                                )
                            
                            # Add data
                            for row_num, row_data in enumerate(boq_display_df.values, 2):
                                for col_num, value in enumerate(row_data, 1):
                                    cell = ws.cell(row=row_num, column=col_num, value=value)
                                    cell.border = Border(
                                        left=Side(style='thin'),
                                        right=Side(style='thin'),
                                        top=Side(style='thin'),
                                        bottom=Side(style='thin')
                                    )
                                    
                                # Color code by status
                                if 'Status' in headers and col_num == headers.index('Status') + 1:
                                    if value == 'approved':
                                        cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                                    elif value == 'pending':
                                        cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
                                    elif value == 'under_review':
                                        cell.fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")
                                    elif value == 'rejected':
                                        cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
                            
                            # Auto-adjust column widths
                            for column in ws.columns:
                                max_length = 0
                                column_letter = column[0].column_letter
                                for cell in column:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except:
                                        pass
                                adjusted_width = min(max_length + 2, 50)
                                ws.column_dimensions[column_letter].width = adjusted_width
                            
                            # Save to bytes
                            excel_buffer = io.BytesIO()
                            wb.save(excel_buffer)
                            excel_buffer.seek(0)
                            
                            st.download_button(
                                label="📈 Export to Excel",
                                data=excel_buffer.getvalue(),
                                file_name=f"BOQ_{selected_project['tracking_id']}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                        except ImportError:
                            st.info("📈 Excel export requires openpyxl package")
                    
                    with col_export3:
                        # Print-friendly view
                        if st.button("🖨️ Print View", use_container_width=True):
                            st.markdown("### 🖨️ Print-Friendly BOQ View")
                            st.markdown(f"**Project:** {selected_project['project_name']} ({selected_project['tracking_id']})")
                            st.markdown(f"**Date:** {datetime.now().strftime('%Y-%m-%d %H:%M')}")
                            st.markdown("---")
                            st.dataframe(boq_display_df, use_container_width=True, hide_index=True)
                    
                    # Project Cost Summary
                    total_project_cost = materials_df['amount_inr'].sum()
                    st.markdown("---")
                    st.markdown("### 💰 Project Cost Summary")
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Total Material Cost", f"₹{total_project_cost:,.2f}")
                    with col2:
                        st.metric("Total Material Items", len(materials_df))
                    with col3:
                        categories_count = len(materials_df['category'].unique())
                        st.metric("Number of Categories", categories_count)
                
                else:
                    st.warning("⚠️ No BOQ/Materials data found for this project.")
                    st.info("This project may have been submitted without material details.")
                    
                    # Show project review actions even without BOQ data
                    st.markdown("#### 📋 Project Review Actions")
                    
                    # Review comments
                    review_comments = st.text_area(
                        "Review Comments",
                        placeholder="Enter your review comments here...",
                        height=100,
                        key=f"review_comments_no_boq_{selected_project['id']}"
                    )
                    
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        if st.button("✅ Approve Project", key=f"approve_no_materials_{selected_project['id']}", use_container_width=True):
                            if review_comments.strip():
                                update_project_status(selected_project['tracking_id'], 'approved', review_comments, 'Admin')
                            else:
                                update_project_status(selected_project['tracking_id'], 'approved', 'Approved by admin', 'Admin')
                            st.success("✅ Project approved!")
                            st.rerun()
                    
                    with col2:
                        if st.button("❌ Reject Project", key=f"reject_no_materials_{selected_project['id']}", use_container_width=True):
                            if review_comments.strip():
                                update_project_status(selected_project['tracking_id'], 'rejected', review_comments, 'Admin')
                            else:
                                update_project_status(selected_project['tracking_id'], 'rejected', 'Rejected by admin', 'Admin')
                            st.error("❌ Project rejected!")
                            st.rerun()
                    
                    with col3:
                        if st.button("🔍 Mark Under Review", key=f"review_no_materials_{selected_project['id']}", use_container_width=True):
                            if review_comments.strip():
                                update_project_status(selected_project['tracking_id'], 'under_review', review_comments, 'Admin')
                            else:
                                update_project_status(selected_project['tracking_id'], 'under_review', 'Marked under review by admin', 'Admin')
                            st.info("🔍 Project marked under review!")
                            st.rerun()
                    
                    with col4:
                        if st.button("📋 Request More Info", key=f"info_no_materials_{selected_project['id']}", use_container_width=True):
                            if review_comments.strip():
                                update_project_status(selected_project['tracking_id'], 'pending', f"More information requested: {review_comments}", 'Admin')
                                st.warning("📋 More information requested!")
                            else:
                                st.warning("Please provide comments when requesting more information.")
            
            st.markdown('</div>', unsafe_allow_html=True)
        
        else:
            st.info("📋 No projects found for review.")
            st.markdown("**To test the admin panel:**")
            st.markdown("1. Go to '📝 Submit Project' page")
            st.markdown("2. Fill in project details and add BOQ materials")
            st.markdown("3. Submit the project")
            st.markdown("4. Return here to review it")
    
    elif admin_password:
        st.error("❌ Invalid admin password.")

# Super User Dashboard
elif page == "👑 Super User Dashboard":
    # Page Header with Breadcrumb
    st.markdown("""
    <div style="margin-bottom: 2rem;">
        <nav style="margin-bottom: 1rem;">
            <span style="color: #6c757d;">🏠 Dashboard</span> 
            <span style="color: #6c757d;">›</span> 
            <span style="color: #2c3e50; font-weight: 600;">Super User Dashboard</span>
        </nav>
        <h1 style="color: #2c3e50; margin: 0;">👑 Super User Dashboard</h1>
        <p style="color: #6c757d; margin: 0.5rem 0 0 0;">Advanced analytics, cumulative calculations, and detailed project insights</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Super User authentication
    super_password = st.text_input("Super User Password", type="password", placeholder="Enter super user password")
    
    if super_password == "super123":  # Super user password
        st.success("✅ Super User access granted!")
        
        # Get comprehensive data
        projects_df = get_projects()
        
        if not projects_df.empty:
            # Cumulative Cost Analysis
            st.markdown('<div class="form-container">', unsafe_allow_html=True)
            st.markdown("### 💰 Cumulative Cost Analysis")
            
            # Get all materials data for cumulative analysis
            conn = sqlite3.connect('project_management.db')
            materials_df = pd.read_sql_query("SELECT * FROM project_materials", conn)
            conn.close()
            
            if not materials_df.empty:
                # Calculate cumulative costs by category
                category_costs = materials_df.groupby('category')['amount_inr'].sum().reset_index()
                category_costs = category_costs.sort_values('amount_inr', ascending=False)
                
                col1, col2 = st.columns(2)
            
            with col1:
                    st.markdown("#### 📊 Cost Distribution by Category")
                    fig_category = px.pie(
                        category_costs,
                        values='amount_inr',
                        names='category',
                        title="Material Cost Distribution",
                        color_discrete_sequence=px.colors.qualitative.Set3
                    )
                    fig_category.update_layout(height=400)
                    st.plotly_chart(fig_category, use_container_width=True)
                
            with col2:
                    st.markdown("#### 📈 Top Categories by Cost")
                    fig_bar = px.bar(
                        category_costs.head(10),
                        x='amount_inr',
                        y='category',
                        orientation='h',
                        title="Top 10 Categories by Cost",
                        color='amount_inr',
                        color_continuous_scale='Blues'
                    )
                    fig_bar.update_layout(height=400)
                    st.plotly_chart(fig_bar, use_container_width=True)
                
                # Detailed cost breakdown
            st.markdown("#### 📋 Detailed Cost Breakdown")
            st.dataframe(category_costs, use_container_width=True)
                
                # Total cumulative costs
            total_material_cost = materials_df['amount_inr'].sum()
            total_projects = len(projects_df)
            avg_cost_per_project = total_material_cost / total_projects if total_projects > 0 else 0
                
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                    st.metric("Total Material Cost", f"₹{total_material_cost:,.2f}")
            with col2:
                    st.metric("Total Projects", total_projects)
            with col3:
                    st.metric("Avg Cost per Project", f"₹{avg_cost_per_project:,.2f}")
            with col4:
                    st.metric("Material Entries", len(materials_df))
            
            st.markdown('</div>', unsafe_allow_html=True)
            
            # Project Performance Analysis
            st.markdown('<div class="form-container">', unsafe_allow_html=True)
            st.markdown("### 📊 Project Performance Analysis")
            
            # Status analysis
            status_analysis = projects_df.groupby('status').agg({
                'id': 'count',
                'total_cost': ['sum', 'mean']
            }).round(2)
            status_analysis.columns = ['Count', 'Total Cost', 'Average Cost']
            
            col1, col2 = st.columns(2)
                    
            with col1:
                st.markdown("#### 📈 Status Distribution")
                fig_status = px.pie(
                    values=projects_df['status'].value_counts().values,
                    names=projects_df['status'].value_counts().index,
                    title="Project Status Distribution",
                    color_discrete_map={
                        'pending': '#ffc107',
                        'approved': '#28a745',
                        'rejected': '#dc3545',
                        'under_review': '#6f42c1'
                    }
                )
                st.plotly_chart(fig_status, use_container_width=True)
                
                with col2:
                    st.markdown("#### 💰 Cost by Status")
                status_costs = projects_df.groupby('status')['total_cost'].sum().reset_index()
                fig_cost = px.bar(
                    status_costs,
                    x='status',
                    y='total_cost',
                    title="Total Cost by Status",
                    color='total_cost',
                    color_continuous_scale='Blues'
                )
                st.plotly_chart(fig_cost, use_container_width=True)
            
            st.markdown("#### 📋 Status Analysis Summary")
            st.dataframe(status_analysis, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)
            
            # Domain Analysis
            st.markdown('<div class="form-container">', unsafe_allow_html=True)
            st.markdown("### 🏗️ Domain Analysis")
            
            domain_analysis = projects_df.groupby('domain').agg({
                'id': 'count',
                'total_cost': ['sum', 'mean']
            }).round(2)
            domain_analysis.columns = ['Project Count', 'Total Cost', 'Average Cost']
            
            col1, col2 = st.columns(2)
                    
            with col1:
                st.markdown("#### 📊 Projects by Domain")
                domain_counts = projects_df['domain'].value_counts().reset_index()
                domain_counts.columns = ['Domain', 'Count']
                fig_domain = px.bar(
                    domain_counts,
                    x='Domain',
                    y='Count',
                    title="Project Count by Domain",
                    color='Count',
                    color_continuous_scale='Blues'
                )
                st.plotly_chart(fig_domain, use_container_width=True)
                    
                with col2:
                    st.markdown("#### 💰 Cost by Domain")
                domain_costs = projects_df.groupby('domain')['total_cost'].sum().reset_index()
                fig_domain_cost = px.pie(
                    domain_costs,
                    values='total_cost',
                    names='domain',
                    title="Cost Distribution by Domain"
                )
                st.plotly_chart(fig_domain_cost, use_container_width=True)
            
            st.markdown("#### 📋 Domain Analysis Summary")
            st.dataframe(domain_analysis, use_container_width=True)
            st.markdown('</div>', unsafe_allow_html=True)
            
            # Export functionality
            st.markdown('<div class="form-container">', unsafe_allow_html=True)
            st.markdown("### 📤 Export Data")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                if st.button("📊 Export Projects Data"):
                    csv = projects_df.to_csv(index=False)
                    st.download_button(
                        label="Download Projects CSV",
                        data=csv,
                        file_name="projects_data.csv",
                        mime="text/csv"
                    )
            
            with col2:
                if st.button("📦 Export Materials Data"):
                    csv = materials_df.to_csv(index=False)
                    st.download_button(
                        label="Download Materials CSV",
                        data=csv,
                        file_name="materials_data.csv",
                        mime="text/csv"
                    )
                    
                    with col3:
                        if st.button("📈 Export Analytics Summary"):
                            summary_data = {
                        'Total Projects': len(projects_df),
                        'Total Material Cost': materials_df['amount_inr'].sum() if not materials_df.empty else 0,
                        'Average Project Cost': projects_df['total_cost'].mean(),
                        'Approval Rate': len(projects_df[projects_df['status'] == 'approved']) / len(projects_df) * 100
                    }
                    summary_df = pd.DataFrame(list(summary_data.items()), columns=['Metric', 'Value'])
                    csv = summary_df.to_csv(index=False)
                    st.download_button(
                        label="Download Summary CSV",
                        data=csv,
                        file_name="analytics_summary.csv",
                        mime="text/csv"
                    )
            
            st.markdown('</div>', unsafe_allow_html=True)
        
        else:
            st.info("No projects found for analysis.")
    
    elif super_password:
        st.error("❌ Invalid super user password.")

# Analytics
elif page == "📊 Analytics":
    st.header("📊 Advanced Analytics")
    
    projects_df = get_projects()
    
    if not projects_df.empty:
        # Time series analysis
        st.subheader("📈 Project Timeline Analysis")
        projects_df['submitted_at'] = pd.to_datetime(projects_df['submitted_at'])
        timeline_df = projects_df.groupby(projects_df['submitted_at'].dt.date).size().reset_index()
        timeline_df.columns = ['Date', 'Projects Submitted']
        
        fig_timeline = px.line(timeline_df, x='Date', y='Projects Submitted', title='Projects Submitted Over Time')
        st.plotly_chart(fig_timeline, use_container_width=True)
        
        # Budget analysis
        st.subheader("💰 Budget Analysis")
        col1, col2 = st.columns(2)
        
        with col1:
            budget_by_domain = projects_df.groupby('domain')['total_cost'].sum().reset_index()
            fig_budget = px.bar(budget_by_domain, x='domain', y='total_cost', title='Total Budget by Domain')
            st.plotly_chart(fig_budget, use_container_width=True)
        
        with col2:
            budget_by_priority = projects_df.groupby('priority')['total_cost'].sum().reset_index()
            fig_priority = px.pie(budget_by_priority, values='total_cost', names='priority', title='Budget Distribution by Priority')
            st.plotly_chart(fig_priority, use_container_width=True)
        
        # Performance metrics
        st.subheader("📊 Performance Metrics")
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            avg_project_cost = projects_df['total_cost'].mean()
            st.metric("Average Project Cost", f"₹{avg_project_cost:,.2f}")
        
        with col2:
            approval_rate = len(projects_df[projects_df['status'] == 'approved']) / len(projects_df) * 100
            st.metric("Approval Rate", f"{approval_rate:.1f}%")
        
        with col3:
            avg_review_time = "N/A"  # Would need more complex calculation
            st.metric("Avg Review Time", avg_review_time)
        
        with col4:
            total_budget = projects_df['total_cost'].sum()
            st.metric("Total Budget", f"₹{total_budget:,.2f}")
    
    else:
        st.info("No data available for analytics.")

# File Upload (Original functionality)
elif page == "📁 File Upload":
    st.header("📁 File Upload & ETL Processing")
    
    # File upload
    uploaded_file = st.file_uploader("Choose a file", type=['csv', 'xlsx', 'xls'])
    
    if uploaded_file:
        try:
            # Save uploaded file
            save_path = os.path.join('data', 'uploads', uploaded_file.name)
            os.makedirs('data/uploads', exist_ok=True)
            
            with open(save_path, 'wb') as f:
                f.write(uploaded_file.getbuffer())
            
            st.success(f"✅ File uploaded: {uploaded_file.name}")
            
            # Process file
            if uploaded_file.name.endswith('.csv'):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)
            
            st.subheader("📊 File Preview")
            st.dataframe(df.head(), use_container_width=True)
            
            # Basic analysis
            st.subheader("📈 Quick Analysis")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.metric("Total Rows", len(df))
            
            with col2:
                st.metric("Total Columns", len(df.columns))
            
            with col3:
                numeric_cols = df.select_dtypes(include=['number']).columns
                st.metric("Numeric Columns", len(numeric_cols))
            
            # ETL processing button
            if st.button("🔄 Run ETL Processing"):
                with st.spinner("Processing file..."):
                    try:
                        # Simple ETL processing
                        processed_df = df.copy()
                        
                        # Clean numeric columns
                        for col in numeric_cols:
                            processed_df[col] = pd.to_numeric(processed_df[col], errors='coerce').fillna(0)
                        
                        # Save processed data
                        processed_path = os.path.join('data', 'processed', f'processed_{uploaded_file.name}')
                        os.makedirs('data/processed', exist_ok=True)
                        processed_df.to_csv(processed_path, index=False)
                        
                        st.success("✅ ETL processing completed!")
                        st.info(f"Processed file saved to: {processed_path}")
                        
                    except Exception as e:
                        st.error(f"❌ ETL processing failed: {str(e)}")
        
        except Exception as e:
            st.error(f"❌ Error processing file: {str(e)}")
        
        # Footer
        st.markdown("---")

# Material Entry Page
elif page == "📦 Material Entry":
    # Page Header with Breadcrumb
    st.markdown("""
    <div style="margin-bottom: 2rem;">
        <nav style="margin-bottom: 1rem;">
            <span style="color: #6c757d;">🏠 Dashboard</span> 
            <span style="color: #6c757d;">›</span> 
            <span style="color: #2c3e50; font-weight: 600;">Material Entry</span>
        </nav>
        <h1 style="color: #2c3e50; margin: 0;">📦 Material Entry Form</h1>
        <p style="color: #6c757d; margin: 0.5rem 0 0 0;">Manage project materials and Bill of Quantities (BOQ)</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Define the categories/topics with their subtopics
    topics_with_subtopics = {
        "1. Supply": [
            "Fabrication (without machining)",
            "Fabrication (with machining)",
            "Raw Materials",
            "Components & Parts",
            "Equipment Supply",
            "Tools & Instruments"
        ],
        "2. Erection": [
            "Structural Erection",
            "Mechanical Erection",
            "Electrical Erection",
            "Piping Erection",
            "Instrumentation Erection",
            "Civil Works"
        ],
        "3. Erection & Commissioning": [
            "Installation & Testing",
            "Commissioning Services",
            "Start-up Support",
            "Performance Testing",
            "Training & Documentation",
            "Warranty Support"
        ],
        "4. Running Expenses": [
            "Fuel & Energy",
            "Maintenance Materials",
            "Spare Parts",
            "Consumables",
            "Utilities",
            "Operational Supplies"
        ],
        "5. Project Management Service PMC": [
            "Project Planning",
            "Quality Control",
            "Safety Management",
            "Progress Monitoring",
            "Coordination Services",
            "Documentation Management"
        ],
        "6. Travels & Others": [
            "Travel Expenses",
            "Accommodation",
            "Transportation",
            "Communication",
            "Miscellaneous Expenses",
            "Contingency"
        ],
        "7. Bonds & Guarantee": [
            "Performance Bond",
            "Bank Guarantee",
            "Warranty Bond",
            "Insurance",
            "Retention Money",
            "Security Deposits"
        ],
        "8. Statutory expenses": [
            "Taxes & Duties",
            "Licenses & Permits",
            "Regulatory Compliance",
            "Environmental Clearances",
            "Safety Certifications",
            "Legal Fees"
        ],
        "9. Over head - Chennai office": [
            "Office Rent",
            "Utilities",
            "Staff Salaries",
            "Administrative Costs",
            "IT Infrastructure",
            "General Expenses"
        ],
        "10. Export": [
            "Export Documentation",
            "Shipping & Logistics",
            "Customs Clearance",
            "International Compliance",
            "Currency Exchange",
            "Export Incentives"
        ],
        "11. Engineering & PMC Support - Third Party": [
            "Design Engineering",
            "Technical Consultancy",
            "Third Party Inspection",
            "Testing Services",
            "Certification Services",
            "Expert Consultation"
        ],
        "12. Service Charges": [
            "Professional Services",
            "Consultancy Fees",
            "Technical Support",
            "Maintenance Services",
            "Training Services",
            "After Sales Support"
        ],
        "13. Royaltee": [
            "Technology License",
            "Patent Fees",
            "Intellectual Property",
            "Software Licenses",
            "Brand Licensing",
            "Technical Know-how"
        ],
        "14. Contigencies": [
            "Project Contingency",
            "Price Escalation",
            "Scope Changes",
            "Risk Mitigation",
            "Unforeseen Events",
            "Buffer Amount"
        ],
        "15. Net Margin": [
            "Profit Margin",
            "Overhead Recovery",
            "Risk Premium",
            "Company Profit",
            "Return on Investment",
            "Financial Returns"
        ]
    }
    
    categories = list(topics_with_subtopics.keys())
    
    units_options = ["kgs", "MT", "mtrs", "Sq.ft", "Sq.Mtr", "Litres", "RMT", "No,s"]
    nos_options = list(range(1, 21))
    source_type_options = ["Vendor Quote", "Company Costing", "Free Issue"]
    payment_schedule_options = ["Ontime", "Monthly"]

    if "materials_data" not in st.session_state:
        st.session_state["materials_data"] = []
    
    # Filter by category option
    if "selected_category_filter" not in st.session_state:
        st.session_state["selected_category_filter"] = "All Categories"
    
    st.markdown('<div class="form-container">', unsafe_allow_html=True)
    st.markdown("### ➕ Add New Material Entry")
    st.markdown('<div class="form-section">', unsafe_allow_html=True)
    
    with st.form("material_entry_form"):
        # First select the category/topic
        category = st.selectbox("Select Category/Topic", categories, key="category_select")
        
        # Then select subtopic based on the selected category
        subtopics = topics_with_subtopics[category]
        subtopic = st.selectbox("Select Sub-topic", subtopics, key="subtopic_select")
        
        # Then enter additional description details
        description = st.text_input("Additional Description (Optional)", placeholder="Add more details if needed")
        
        col1, col2, col3 = st.columns([2, 1, 1])
        with col1:
            units = st.selectbox("Units/Qty", units_options)
        with col2:
            nos = st.selectbox("Nos", nos_options)
        with col3:
            unit_price = st.number_input("Unit Price Total Amount", min_value=0.0, step=0.01, format="%.2f")
        
        col1, col2 = st.columns(2)
        with col1:
            source_type = st.selectbox("Source/Type", source_type_options)
        with col2:
            payment_schedule = st.selectbox("Payment Schedule", payment_schedule_options)
        
        amount_inr = nos * unit_price
        st.markdown(f"##### Amount INR: <span style='color: #3973ac; font-size: 22px;'>₹{amount_inr:,.2f}</span>", unsafe_allow_html=True)
        
        # Justification section
        st.markdown("#### 📄 Justification for this Material Entry")
        justification_type = st.radio("Justification Type:", ["Text Input", "File Upload"], horizontal=True, key="mat_justification_type")
        
        if justification_type == "Text Input":
            material_justification = st.text_area("Justification *", height=80, placeholder="Provide justification for this material entry...", key="mat_justification_text")
        else:
            material_justification_file = st.file_uploader("Upload Justification Document *", type=['pdf', 'docx', 'txt'], help="Upload PDF, DOCX, or TXT file", key="mat_justification_file")
            material_justification = material_justification_file.name if material_justification_file else ""
        
        submit = st.form_submit_button("Add Entry")

    if submit:
        # Combine subtopic and additional description
        full_description = f"{subtopic}"
        if description:
            full_description += f" - {description}"
        
        # Handle justification
        justification_content = material_justification
        justification_file_path = ""
        
        if justification_type == "File Upload" and material_justification_file:
            # Save uploaded file
            upload_dir = "data/uploads/material_justifications"
            os.makedirs(upload_dir, exist_ok=True)
            file_path = os.path.join(upload_dir, material_justification_file.name)
            with open(file_path, "wb") as f:
                f.write(material_justification_file.getbuffer())
            justification_file_path = file_path
            justification_content = f"File uploaded: {material_justification_file.name}"
        
        # Validate required fields
        required_fields = [subtopic, units, nos, source_type, payment_schedule]
        if justification_type == "Text Input":
            required_fields.append(material_justification)
        else:
            required_fields.append(material_justification_file)
        
        if all(required_fields) and unit_price > 0:
            entry = {
            "Category": category,
                "Sub-topic": subtopic,
                "Description": full_description,
            "Units/Qty": units,
            "Nos": nos,
            "Source/Type": source_type,
            "Payment Schedule": payment_schedule,
            "Unit Price Total Amount": unit_price,
                "Amount INR": amount_inr,
                "Justification": justification_content,
                "Justification Type": justification_type,
                "Justification File Path": justification_file_path
        }
        st.session_state["materials_data"].append(entry)
        st.success(f"Entry added to category: {category} - {subtopic}")
    else:
        st.error("❌ Please fill all required fields and enter Unit Price > 0.")

    # Display entries with filtering option
    if st.session_state["materials_data"]:
        st.markdown("---")
        st.subheader("Your Material Entries")
        
        # Add category filter
        filter_options = ["All Categories"] + categories
        selected_filter = st.selectbox(
            "Filter by Category", 
            filter_options, 
            index=filter_options.index(st.session_state["selected_category_filter"])
        )
        st.session_state["selected_category_filter"] = selected_filter
        
        # Filter the data based on selection
        if selected_filter == "All Categories":
            filtered_data = st.session_state["materials_data"]
        else:
            filtered_data = [item for item in st.session_state["materials_data"] if item["Category"] == selected_filter]
        
        if filtered_data:
            # Group by category for display
            categories_in_data = sorted(set(item["Category"] for item in filtered_data))
            
            for cat in categories_in_data:
                with st.expander(f"{cat}", expanded=True):
                    cat_data = [item for item in filtered_data if item["Category"] == cat]
                    cat_df = pd.DataFrame(cat_data)
                    
                    # Calculate totals for this category
                    total_amount = sum(item["Amount INR"] for item in cat_data)
                    
                    # Display the data table with subtopic information
                    display_columns = ["Sub-topic", "Description", "Units/Qty", "Nos", "Unit Price Total Amount", "Amount INR"]
                    st.dataframe(
                        cat_df[display_columns], 
                        use_container_width=True
                    )
                    
                    st.markdown(f"**Total for {cat}: ₹{total_amount:,.2f}**")
                    
                    # Add delete buttons for each entry in this category
                    if st.button(f"Delete All Entries in {cat}", key=f"del_cat_{cat}"):
                        st.session_state["materials_data"] = [
                            item for item in st.session_state["materials_data"] if item["Category"] != cat
                        ]
                        st.success(f"All entries in {cat} deleted!")
                        st.experimental_rerun()
            
            # Show grand total
            grand_total = sum(item["Amount INR"] for item in filtered_data)
            st.markdown(f"### Grand Total: ₹{grand_total:,.2f}")
            
            # Export options
            if st.button("Export to CSV"):
                export_df = pd.DataFrame(st.session_state["materials_data"])
                csv = export_df.to_csv(index=False)
                st.download_button(
                    label="Download CSV",
                    data=csv,
                    file_name="material_entries.csv",
                    mime="text/csv"
                )
        else:
            st.info(f"No entries found for the selected category: {selected_filter}")
